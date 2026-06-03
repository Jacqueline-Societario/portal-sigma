"""
scripts/atualizar_base_cnae_concla.py
======================================
Detecta automaticamente a versao mais recente da CNAE-Subclasses publicada
no CONCLA/IBGE, baixa o XLSX oficial, converte para JSON e salva em
static/data/cnae_subclasses.json.

Uso:
    python scripts/atualizar_base_cnae_concla.py           # verifica e atualiza
    python scripts/atualizar_base_cnae_concla.py --check   # so verifica versao
    python scripts/atualizar_base_cnae_concla.py --force   # forca atualizacao

Fonte oficial:
    https://concla.ibge.gov.br/classificacoes/download-concla.html
"""

import os
import re
import sys
import json
import shutil
import unicodedata
import tempfile
from datetime import datetime, timezone

import httpx
import openpyxl

# ── Caminhos ──────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR   = os.path.dirname(SCRIPT_DIR)
DATA_DIR   = os.path.join(BASE_DIR, 'static', 'data')
JSON_OUT   = os.path.join(DATA_DIR, 'cnae_subclasses.json')

# ── URLs oficiais ─────────────────────────────────────────────────────────────
CONCLA_PAGINA    = 'https://concla.ibge.gov.br/classificacoes/por-tema/atividades-economicas'
CONCLA_DOWNLOADS = 'https://concla.ibge.gov.br/classificacoes/download-concla.html'
CONCLA_BASE      = 'https://concla.ibge.gov.br'

# Padrao de nome de arquivo: CNAE_Subclasses_X_Y_Estrutura_Detalhada.xlsx
PATTERN_XLSX = re.compile(
    r'href=["\']([^"\']*CNAE_Subclasses_(\d+)_(\d+)_Estrutura_Detalhada\.xlsx)["\']',
    re.IGNORECASE,
)


# ── Normalizacao ──────────────────────────────────────────────────────────────
def normalizar(texto: str) -> str:
    """Remove acentos e converte para minusculo."""
    return unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode().lower()


# ── Deteccao de versao ────────────────────────────────────────────────────────
def detectar_versao_mais_recente() -> tuple:
    """
    Acessa a pagina de downloads do CONCLA e retorna
    (versao_str, url_xlsx, nome_arquivo).

    Ex: ('2.3', 'https://concla.ibge.gov.br/images/.../CNAE_Subclasses_2_3...xlsx',
         'CNAE_Subclasses_2_3_Estrutura_Detalhada.xlsx')
    """
    print(f"[CONCLA] Acessando pagina de downloads: {CONCLA_DOWNLOADS}")
    r = httpx.get(CONCLA_DOWNLOADS, timeout=30, follow_redirects=True)
    r.raise_for_status()

    candidatos = []
    for m in PATTERN_XLSX.finditer(r.text):
        path, major, minor = m.group(1), int(m.group(2)), int(m.group(3))
        url   = path if path.startswith('http') else CONCLA_BASE + path
        nome  = os.path.basename(path.split('?')[0])
        versao_str = f"{major}.{minor}"
        candidatos.append(((major, minor), versao_str, url, nome))

    if not candidatos:
        raise RuntimeError(
            "Nenhum arquivo CNAE_Subclasses_*_Estrutura_Detalhada.xlsx encontrado "
            f"em {CONCLA_DOWNLOADS}. Verifique manualmente se o CONCLA alterou a estrutura da pagina."
        )

    # Ordenar por versao numerica decrescente, pegar a maior
    candidatos.sort(key=lambda x: x[0], reverse=True)
    _, versao_str, url, nome = candidatos[0]

    print(f"[CONCLA] Versoes encontradas: {[c[1] for c in candidatos]}")
    print(f"[CONCLA] Versao mais recente: {versao_str} → {nome}")
    return versao_str, url, nome


# ── Versao instalada ─────────────────────────────────────────────────────────
def versao_instalada() -> str:
    """Retorna a versao registrada no JSON atual, ou '' se inexistente."""
    if not os.path.isfile(JSON_OUT):
        return ''
    try:
        with open(JSON_OUT, encoding='utf-8') as f:
            meta = json.load(f)
        return meta.get('versao_cnae', '')
    except Exception:
        return ''


# ── Download ──────────────────────────────────────────────────────────────────
def baixar_xlsx(url: str, destino: str):
    print(f"[DOWNLOAD] {url}")
    r = httpx.get(url, timeout=60, follow_redirects=True)
    r.raise_for_status()
    if len(r.content) < 10_000:
        raise RuntimeError(
            f"Arquivo muito pequeno ({len(r.content)} bytes). "
            "Provavel erro de download ou redirecionamento inesperado."
        )
    with open(destino, 'wb') as f:
        f.write(r.content)
    print(f"[DOWNLOAD] Salvo: {destino} ({len(r.content):,} bytes)")


# ── Conversao XLSX → registros ────────────────────────────────────────────────
def converter_xlsx(caminho_xlsx: str) -> list:
    """
    Le o XLSX do CONCLA e retorna lista de dicts, um por subclasse.

    Estrutura do XLSX (colunas A-G):
        A: Secao  (ex: 'A')
        B: Divisao (ex: '01')
        C: Grupo   (ex: '01.1')
        D: Classe  (ex: '01.11-3')
        E: Subclasse (ex: '0111-3/01')  ← linha de dados
        F: Denominacao/descricao
        G: (ignorada nesta versao)
    """
    wb = openpyxl.load_workbook(caminho_xlsx, read_only=True, data_only=True)

    # Localizar aba com estrutura detalhada
    aba = None
    for nome in wb.sheetnames:
        n = nome.lower()
        if 'estrutura' in n or ('subclass' in n and 'resumo' not in n and 'preferen' not in n):
            aba = wb[nome]
            break
    if aba is None:
        aba = wb[wb.sheetnames[0]]
    print(f"[XLSX] Aba selecionada: '{aba.title}'")

    subclasses = []
    secao_cod = secao_desc = ''
    divisao_cod = divisao_desc = ''
    grupo_cod = grupo_desc = ''
    classe_cod = classe_desc = ''

    for row in aba.iter_rows(values_only=True):
        if len(row) < 6:
            continue
        col_sec, col_div, col_grp, col_cls, col_sub, col_desc = (
            row[0], row[1], row[2], row[3], row[4],
            row[5] if row[5] is not None else ''
        )
        desc_str = str(col_desc).strip() if col_desc else ''

        # Secao
        if col_sec and not col_div and not col_grp and not col_cls and not col_sub:
            secao_cod = str(col_sec).strip()
            secao_desc = desc_str
            divisao_cod = divisao_desc = grupo_cod = grupo_desc = classe_cod = classe_desc = ''
            continue

        # Divisao
        if col_div and not col_grp and not col_cls and not col_sub:
            divisao_cod = str(col_div).strip()
            divisao_desc = desc_str
            grupo_cod = grupo_desc = classe_cod = classe_desc = ''
            continue

        # Grupo
        if col_grp and not col_cls and not col_sub:
            grupo_cod = str(col_grp).strip()
            grupo_desc = desc_str
            classe_cod = classe_desc = ''
            continue

        # Classe
        if col_cls and not col_sub:
            classe_cod = str(col_cls).strip()
            classe_desc = desc_str
            continue

        # Subclasse — linha de dados real
        if col_sub:
            cod_fmt = str(col_sub).strip()
            cod_sem = re.sub(r'\D', '', cod_fmt)

            # termos_normalizados: descricao + contexto hierarquico (sem acentos)
            termos = normalizar(
                f"{desc_str} {classe_desc} {grupo_desc} {divisao_desc} {secao_desc}"
            )

            subclasses.append({
                'codigo_sem_mascara': cod_sem,
                'codigo_formatado':   cod_fmt,
                'descricao':          desc_str,
                'secao':              secao_cod,
                'secao_desc':         secao_desc,
                'divisao':            divisao_cod,
                'divisao_desc':       divisao_desc,
                'grupo':              grupo_cod,
                'grupo_desc':         grupo_desc,
                'classe':             classe_cod,
                'classe_desc':        classe_desc,
                'termos_normalizados': termos,
            })

    return subclasses


# ── Salvar JSON ───────────────────────────────────────────────────────────────
def salvar_json(subclasses: list, versao: str, url_xlsx: str, nome_xlsx: str):
    os.makedirs(DATA_DIR, exist_ok=True)

    # Backup da base anterior
    if os.path.isfile(JSON_OUT):
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        bak = JSON_OUT.replace('.json', f'.bak_{ts}.json')
        shutil.copy2(JSON_OUT, bak)
        print(f"[BACKUP] Base anterior salva em: {bak}")

    agora = datetime.now(timezone.utc).isoformat()
    payload = {
        'fonte':                'CONCLA/IBGE',
        'pagina_oficial':       CONCLA_PAGINA,
        'pagina_downloads':     CONCLA_DOWNLOADS,
        'versao_cnae':          versao,
        'arquivo_origem':       nome_xlsx,
        'url_arquivo_origem':   url_xlsx,
        'data_download':        agora,
        'data_geracao_json':    agora,
        'quantidade_registros': len(subclasses),
        'subclasses':           subclasses,
    }

    tmp = JSON_OUT + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, separators=(',', ':'))

    # Validar antes de substituir
    with open(tmp, encoding='utf-8') as f:
        check = json.load(f)
    if len(check.get('subclasses', [])) < 100:
        os.remove(tmp)
        raise RuntimeError("JSON gerado tem menos de 100 registros — abortando substituicao.")

    shutil.move(tmp, JSON_OUT)
    print(f"[JSON] Salvo: {JSON_OUT} ({len(subclasses)} registros)")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    modo_check = '--check' in sys.argv
    modo_force = '--force' in sys.argv

    print("=" * 60)
    print("ATUALIZADOR DE BASE CNAE — CONCLA/IBGE")
    print("=" * 60)

    # Versao instalada
    v_atual = versao_instalada()
    if v_atual:
        print(f"[LOCAL] Versao instalada: {v_atual}")
    else:
        print("[LOCAL] Nenhuma base instalada.")

    # Versao disponivel no CONCLA
    try:
        v_remota, url_xlsx, nome_xlsx = detectar_versao_mais_recente()
    except Exception as e:
        print(f"[ERRO] Nao foi possivel acessar o CONCLA: {e}")
        sys.exit(1)

    # Comparar versoes
    def parse_v(s):
        try:
            return tuple(int(x) for x in str(s).split('.'))
        except Exception:
            return (0,)

    v_atual_t  = parse_v(v_atual)
    v_remota_t = parse_v(v_remota)

    if v_atual and v_remota_t <= v_atual_t and not modo_force:
        print(f"[OK] Base local ({v_atual}) ja e a mais recente. Nenhuma acao necessaria.")
        print("     Use --force para forcar atualizacao mesmo assim.")
        sys.exit(0)

    if modo_check:
        if v_remota_t > v_atual_t:
            print(f"[NOVA VERSAO] Disponivel: {v_remota} (instalada: {v_atual or 'nenhuma'})")
            print(f"              Execute sem --check para atualizar.")
        sys.exit(0)

    if v_remota_t > v_atual_t:
        print(f"[ATUALIZACAO] {v_atual or 'nenhuma'} → {v_remota}")
    else:
        print(f"[FORCE] Reprocessando versao {v_remota}...")

    # Download em arquivo temporario
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = tmp.name
    try:
        baixar_xlsx(url_xlsx, tmp_path)

        # Converter
        print("[CONVERSAO] Processando XLSX...")
        subclasses = converter_xlsx(tmp_path)
        print(f"[CONVERSAO] {len(subclasses)} subclasses extraidas")

        if len(subclasses) < 100:
            raise RuntimeError(f"Apenas {len(subclasses)} registros convertidos — arquivo invalido.")

        # Salvar JSON
        salvar_json(subclasses, v_remota, url_xlsx, nome_xlsx)

    except Exception as e:
        print(f"[ERRO] {e}")
        print("[SEGURANCA] Base anterior preservada. Nenhuma alteracao realizada.")
        sys.exit(1)
    finally:
        if os.path.isfile(tmp_path):
            os.remove(tmp_path)

    print()
    print("=" * 60)
    print(f"[SUCESSO] Base CNAE {v_remota} instalada com {len(subclasses)} subclasses.")
    print(f"          Arquivo: {JSON_OUT}")
    print("=" * 60)


if __name__ == '__main__':
    main()
