"""
scripts/gerar_base_cnae_v2.py
==============================
Gera a base CNAE enriquecida (formato 2) em static/data/cnae_subclasses.json,
consolidando quatro fontes:

  1. Estrutura oficial   — static/data/cnae_subclasses.json atual (planilha CONCLA)
  2. Notas explicativas  — API oficial servicodados.ibge.gov.br (v2/cnae/subclasses)
  3. Descritores         — static/data/descritores_concla_raw.json (raspagem CONCLA,
                           gerado por scripts/raspar_descritores_concla.py)
  4. Correspondencias    — aba "RESUMO ALTERACOES" da planilha oficial 2.3
                           (mapa DE 2.2 -> PARA 2.3)

Uso:
    python scripts/gerar_base_cnae_v2.py --xlsx /caminho/CNAE_Subclasses_2_3_Estrutura_Detalhada.xlsx
    python scripts/gerar_base_cnae_v2.py          # baixa a planilha do CONCLA

Seguranca: valida contagens e cobertura ANTES de substituir; backup .bak_
timestamped; escrita atomica. Se qualquer validacao falhar, a base atual
permanece intacta. Campos do formato 1 sao preservados (retrocompativel).
"""
import os
import re
import sys
import json
import shutil
import tempfile
from datetime import datetime, timezone

import httpx
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR   = os.path.dirname(SCRIPT_DIR)
DATA_DIR   = os.path.join(BASE_DIR, 'static', 'data')
BASE_JSON  = os.path.join(DATA_DIR, 'cnae_subclasses.json')
CACHE_DESC = os.path.join(DATA_DIR, 'descritores_concla_raw.json')

API_SUBCLASSES = 'https://servicodados.ibge.gov.br/api/v2/cnae/subclasses'

COBERTURA_MINIMA_DESC  = 0.95
COBERTURA_MINIMA_NOTAS = 0.95
TOTAL_MINIMO           = 1300

RE_COD_FMT = re.compile(r'^\d{4}-\d/\d{2}$')


def limpar_notas(observacoes: list) -> str:
    """Junta as observacoes da API em texto corrido limpo."""
    texto = ' '.join(o for o in (observacoes or []) if o)
    texto = texto.replace('\r\n', ' ').replace('\n', ' ').replace('#', ' ')
    return ' '.join(texto.split())


def extrair_correspondencias(caminho_xlsx: str) -> list:
    """Le a aba RESUMO ALTERACOES e devolve pares DE(2.2) -> PARA(2.3)."""
    wb = openpyxl.load_workbook(caminho_xlsx, read_only=True, data_only=True)
    aba = None
    for nome in wb.sheetnames:
        if 'RESUMO' in nome.upper():
            aba = wb[nome]
            break
    if aba is None:
        print("[AVISO] Aba RESUMO ALTERACOES nao encontrada — correspondencias vazias.")
        return []

    pares = []
    de_cod = de_desc = ''
    for row in aba.iter_rows(values_only=True):
        vals = [str(v).strip() if v is not None else '' for v in row]
        vals += [''] * (7 - len(vals))
        col_de, col_de_desc, col_para, col_para_desc, col_obs = (
            vals[1], vals[2], vals[3], vals[5], vals[6])
        if RE_COD_FMT.match(col_de):
            de_cod, de_desc = col_de, col_de_desc
        if RE_COD_FMT.match(col_para) and de_cod:
            pares.append({
                'de': de_cod, 'de_descricao': de_desc,
                'para': col_para, 'para_descricao': col_para_desc,
                'observacao': col_obs,
            })
    return pares


def main():
    xlsx_path = None
    if '--xlsx' in sys.argv:
        xlsx_path = sys.argv[sys.argv.index('--xlsx') + 1]

    # ── 1. Estrutura atual ────────────────────────────────────────────────────
    with open(BASE_JSON, encoding='utf-8') as f:
        base = json.load(f)
    meta_v1 = {k: v for k, v in base.items() if k != 'subclasses'}
    subclasses = [s for s in base['subclasses']
                  if re.fullmatch(r'\d{7}', s.get('codigo_sem_mascara', ''))]
    descartados = len(base['subclasses']) - len(subclasses)
    print(f"[ESTRUTURA] {len(subclasses)} subclasses validas "
          f"({descartados} registro(s) invalido(s) descartado(s))")

    # ── 2. Descritores (cache da raspagem) ────────────────────────────────────
    with open(CACHE_DESC, encoding='utf-8') as f:
        cache = json.load(f)
    desc_map = {k: v['descritores'] for k, v in cache['subclasses'].items()}
    print(f"[DESCRITORES] cache com {len(desc_map)} subclasses")

    # ── 3. Notas explicativas (API oficial) ───────────────────────────────────
    print(f"[API] Baixando notas explicativas: {API_SUBCLASSES}")
    r = httpx.get(API_SUBCLASSES, timeout=60, follow_redirects=True)
    r.raise_for_status()
    api = r.json()
    notas_map = {d['id']: limpar_notas(d.get('observacoes')) for d in api}
    print(f"[API] {len(notas_map)} subclasses na API")

    # Diferencas planilha x API (informativo)
    cod_planilha = {s['codigo_sem_mascara'] for s in subclasses}
    so_api      = sorted(set(notas_map) - cod_planilha)
    so_planilha = sorted(cod_planilha - set(notas_map))
    if so_api:
        print(f"[DIFF] Na API e nao na planilha ({len(so_api)}): {so_api[:10]}")
    if so_planilha:
        print(f"[DIFF] Na planilha e nao na API ({len(so_planilha)}): {so_planilha[:10]}")

    # Uniao das fontes oficiais: subclasses presentes so na API entram na base
    # (ex.: 9900-8/00 — a planilha 2.3 para na classe 99.00-8 e omite a subclasse).
    # No proximo refresh o raspador captura os descritores delas automaticamente.
    if so_api:
        api_map = {d['id']: d for d in api}
        for cod in so_api:
            d = api_map[cod]
            classe  = d.get('classe') or {}
            grupo   = classe.get('grupo') or {}
            divisao = grupo.get('divisao') or {}
            secao   = divisao.get('secao') or {}
            desc_api = (d.get('descricao') or '').strip()
            subclasses.append({
                'codigo_sem_mascara': cod,
                'codigo_formatado':   f"{cod[:4]}-{cod[4]}/{cod[5:]}",
                'descricao':          desc_api.capitalize(),
                'secao':              secao.get('id', ''),
                'secao_desc':         secao.get('descricao', ''),
                'divisao':            divisao.get('id', ''),
                'divisao_desc':       divisao.get('descricao', ''),
                'grupo':              grupo.get('id', ''),
                'grupo_desc':         grupo.get('descricao', ''),
                'classe':             classe.get('id', ''),
                'classe_desc':        classe.get('descricao', ''),
                'termos_normalizados': '',
                'origem':             'api_v2',
            })
        print(f"[COMPLEMENTO] {len(so_api)} subclasse(s) da API adicionada(s) a base")

    # ── 4. Correspondencias 2.2 -> 2.3 ────────────────────────────────────────
    if not xlsx_path or not os.path.isfile(xlsx_path):
        sys.path.insert(0, SCRIPT_DIR)
        from atualizar_base_cnae_concla import detectar_versao_mais_recente, baixar_xlsx
        _, url_xlsx, _ = detectar_versao_mais_recente()
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp.close()
        baixar_xlsx(url_xlsx, tmp.name)
        xlsx_path = tmp.name
    correspondencias = extrair_correspondencias(xlsx_path)
    print(f"[CORRESPONDENCIAS] {len(correspondencias)} pares 2.2 -> 2.3")

    # ── 5. Consolidar ─────────────────────────────────────────────────────────
    sem_desc, sem_notas = [], []
    for s in subclasses:
        cod = s['codigo_sem_mascara']
        s['descritores'] = desc_map.get(cod, [])
        s['notas']       = notas_map.get(cod, '')
        if not s['descritores']:
            sem_desc.append(cod)
        if not s['notas']:
            sem_notas.append(cod)

    total = len(subclasses)
    cob_desc  = 1 - len(sem_desc) / total
    cob_notas = 1 - len(sem_notas) / total
    total_descritores = sum(len(s['descritores']) for s in subclasses)
    print(f"[COBERTURA] descritores: {cob_desc:.1%} ({len(sem_desc)} sem) | "
          f"notas: {cob_notas:.1%} ({len(sem_notas)} sem) | "
          f"{total_descritores} descritores no total")

    # ── 6. Validacoes (antes de substituir) ───────────────────────────────────
    erros = []
    if total < TOTAL_MINIMO:
        erros.append(f"total de subclasses {total} < {TOTAL_MINIMO}")
    if cob_desc < COBERTURA_MINIMA_DESC:
        erros.append(f"cobertura de descritores {cob_desc:.1%} < {COBERTURA_MINIMA_DESC:.0%} "
                     f"(faltam: {sem_desc[:10]}...)")
    if cob_notas < COBERTURA_MINIMA_NOTAS:
        erros.append(f"cobertura de notas {cob_notas:.1%} < {COBERTURA_MINIMA_NOTAS:.0%}")
    if erros:
        print("[ERRO] Validacao falhou — base atual PRESERVADA:")
        for e in erros:
            print(f"   - {e}")
        sys.exit(1)

    # ── 7. Montar e salvar ────────────────────────────────────────────────────
    agora = datetime.now(timezone.utc).isoformat()
    payload = dict(meta_v1)
    payload.update({
        'formato': 2,
        'data_geracao_json': agora,
        'quantidade_registros': total,
        'quantidade_descritores': total_descritores,
        'fonte_notas': {'url': API_SUBCLASSES, 'data': agora,
                        'total_api': len(notas_map)},
        'fonte_descritores': {
            'origem': cache.get('fonte', ''),
            'atualizado_em': cache.get('atualizado_em', ''),
            'total_capturado': cache.get('total_capturado', 0)},
        'arquivo_sinonimos': 'static/data/cnae_sinonimos.json',
        'correspondencias_2_2_para_2_3': correspondencias,
        'subclasses': subclasses,
    })

    if os.path.isfile(BASE_JSON):
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        bak = BASE_JSON.replace('.json', f'.bak_{ts}.json')
        shutil.copy2(BASE_JSON, bak)
        print(f"[BACKUP] {bak}")

    tmp_out = BASE_JSON + '.tmp'
    with open(tmp_out, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, separators=(',', ':'))
    with open(tmp_out, encoding='utf-8') as f:
        check = json.load(f)
    if len(check.get('subclasses', [])) < TOTAL_MINIMO:
        os.remove(tmp_out)
        raise RuntimeError("Releitura do JSON gerado falhou — abortado.")
    shutil.move(tmp_out, BASE_JSON)

    tam = os.path.getsize(BASE_JSON)
    print(f"[SUCESSO] Base formato 2 gravada: {BASE_JSON}")
    print(f"          {total} subclasses | {total_descritores} descritores | "
          f"{tam/1024/1024:.1f} MB")


if __name__ == '__main__':
    main()
